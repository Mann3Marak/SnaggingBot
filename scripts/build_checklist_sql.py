from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
import re

ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = ROOT.parent
OUT = ROOT / 'supabase' / 'migrations' / '20250922150000_checklists_from_excel.sql'

FILES = [
    ('T2', REPO_ROOT / 'Snag list Inspection - Blank  - T2.xlsx'),
    ('T2+1', REPO_ROOT / 'Snag list Inspection - Blank  - T2+1.xlsx'),
    ('T3', REPO_ROOT / 'Snag list Inspection - Blank  - T3.xlsx'),
    ('T3+1', REPO_ROOT / 'Snag list Inspection - Blank  - T3+1.xlsx'),
]

# Heuristics: assume first sheet, rows with two language columns or single description
# We will scan each sheet and collect non-empty strings as item_description (EN)
# If a cell contains ' / ' or '-' separators, we keep as description; Portuguese not guaranteed in source.
# The PO requested to use exactly the items present; we will treat each non-empty row in col A as English item,
# and if col B exists and has text, use as Portuguese translation.

def normalize(text: str) -> str:
    text = re.sub(r"\s+", " ", text.strip())
    return text

rows = []
for apt_type, path in FILES:
    if not path.exists():
        print(f"Missing: {path}")
        continue
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    room = None
    order = 0
    # Try detect room headers by bold or uppercase; fallback when cell ends with ':'
    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row if c is not None and str(c).strip() != '']
        if not cells:
            continue
        # header/room heuristics
        first = str(cells[0]).strip()
        if len(cells) == 1 and (first.endswith(':') or first.isupper() or first.lower() in {'kitchen','bathroom','living room','master bedroom','second bedroom','suite','hallway','laundry','study','office','balcony','entrance','entrada','cozinha','casa de banho','sala','quarto','corredor'}):
            room = normalize(first).rstrip(':')
            continue
        # item row, take col A as EN, col B as PT if present
        en = normalize(str(cells[0]))
        pt = normalize(str(cells[1])) if len(cells) > 1 else None
        order += 1
        rows.append((apt_type, room or 'General', en, pt))

# Build SQL with parameterized inserts into checklist_templates avoiding duplicates by unique key we created
sql_lines = [
    "-- Generated from Excel files; using only provided items",
]
for i, (apt, room, en, pt) in enumerate(rows, start=1):
    en_esc = en.replace("'", "''")
    pt_esc = (pt or '').replace("'", "''")
    notes = ''
    sql_lines.append(
        f"insert into checklist_templates (apartment_type, room_type, item_description, item_description_pt, order_sequence) values ('{apt}', '{room.replace("'","''")}', '{en_esc}', {"'"+pt_esc+"'" if pt else 'null'}, {i}) on conflict (apartment_type, room_type, order_sequence) do update set item_description=excluded.item_description, item_description_pt=excluded.item_description_pt;"
    )

OUT.write_text("\n".join(sql_lines), encoding='utf-8')
print(f"Wrote {OUT}")
