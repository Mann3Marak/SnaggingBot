from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
import re

REPO = Path(__file__).resolve().parents[1].parent
APP = Path(__file__).resolve().parents[1]
OUT = APP / 'supabase' / 'migrations' / '20250922161500_checklists_rebuild.sql'

FILES = [
    ('T2', REPO / 'Snag list Inspection - Blank  - T2.xlsx'),
    ('T2+1', REPO / 'Snag list Inspection - Blank  - T2+1.xlsx'),
    ('T3', REPO / 'Snag list Inspection - Blank  - T3.xlsx'),
    ('T3+1', REPO / 'Snag list Inspection - Blank  - T3+1.xlsx'),
]

# Room headings supplied by PO (normalized)
ROOMS_CANON = [
    'Hall',
    'Kitchen',
    'Dining Room',
    'Living Room',
    'Balcony - Pool side',
    'Suite Bedroom',
    'Suite Bathroom',
    'Bedroom 2',
    'Bathroom 2',
    'Balcony - Back',
    'Outside Patio Storage',
]

# Normalize helpers
_ws = re.compile(r"\s+")
_dash = re.compile(r"\s*[–—-]\s*")

def normalize_text(s: str) -> str:
    s = s.replace('\u00a0', ' ')
    s = _dash.sub(' - ', s)
    s = _ws.sub(' ', s).strip()
    # strip trailing ':'
    if s.endswith(':'):
        s = s[:-1].strip()
    return s

ROOMS_SET = set(r.lower() for r in ROOMS_CANON)

rows = []  # (apt_type, room, item)

for apt, path in FILES:
    if not path.exists():
        raise FileNotFoundError(f"Checklist source missing: {path}")
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    current_room = None
    # iterate lines
    for xlrow in ws.iter_rows(values_only=True):
        # collapse row to first two textual cells
        vals = [v for v in xlrow if v not in (None, '')]
        if not vals:
            continue
        # join if many cells contain text fragments
        texts = [normalize_text(str(v)) for v in vals if str(v).strip()]
        if not texts:
            continue
        # Consider single-cell lines as potential room headers
        if len(texts) == 1:
            t = texts[0]
            t_norm = t.lower()
            if t_norm in ROOMS_SET:
                current_room = next(rc for rc in ROOMS_CANON if rc.lower() == t_norm)
                continue
        # Otherwise treat as item(s). We only take the first cell as item description (EN-only for now)
        item_en = texts[0]
        if current_room is None:
            # If no room encountered yet, try to infer from common starts
            current_room = 'Hall'
        rows.append((apt, current_room, item_en))

# Build SQL: wipe old rows per apartment_type then insert
sql = []
sql.append('-- Rebuild checklist_templates from provided Excel files (EN only, rooms normalized)')
sql.append('begin;')
apt_types = {apt for apt,_,_ in rows}
sql.append("delete from checklist_templates where apartment_type in (" + ",".join(["'"+a+"'" for a in sorted(apt_types)]) + ");")

# Keep order per apartment type
order_map: dict[str,int] = {}
for apt, room, item in rows:
    order_map.setdefault(apt, 0)
    order_map[apt] += 1
    i = order_map[apt]
    item_sql = item.replace("'", "''")
    room_sql = room.replace("'", "''")
    sql.append(
        f"insert into checklist_templates (apartment_type, room_type, item_description, item_description_pt, order_sequence) "
        f"values ('{apt}', '{room_sql}', '{item_sql}', null, {i});"
    )

sql.append('commit;')
OUT.write_text("\n".join(sql) + "\n", encoding='utf-8')
print(f"Wrote {OUT}")
