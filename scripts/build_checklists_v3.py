from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
import re, time

REPO = Path(__file__).resolve().parents[1].parent
APP = Path(__file__).resolve().parents[1]
MIG_DIR = APP / 'supabase' / 'migrations'

FILES = [
    ('T2', REPO / 'Snag list Inspection - Blank  - T2.xlsx'),
    ('T2+1', REPO / 'Snag list Inspection - Blank  - T2+1.xlsx'),
    ('T3', REPO / 'Snag list Inspection - Blank  - T3.xlsx'),
    ('T3+1', REPO / 'Snag list Inspection - Blank  - T3+1.xlsx'),
]

# Canonical English room names
ROOMS_CANON = [
    'Hall',
    'Kitchen',
    'Dining Room',
    'Living Room',
    'Balcony - Pool side',
    'Suite Bedroom',
    'Suite Bathroom',
    'Bedroom 2',
    'Bathroom 2',
    'Balcony - Back',
    'Outside Patio Storage',
]

# Portuguese -> English mapping (normalized)
PT_TO_EN = {
    'entrada': 'Hall',
    'cozinha': 'Kitchen',
    'sala de estar': 'Living Room',
    'quarto suite': 'Suite Bedroom',
    'casa de banho suite': 'Suite Bathroom',
    'quarto 2': 'Bedroom 2',
    'casa de banho 2': 'Bathroom 2',
    'varanda - pool side': 'Balcony - Pool side',
    'varanda - traseira': 'Balcony - Back',
    'balcony - pool side': 'Balcony - Pool side',
    'balcony - back': 'Balcony - Back',
    'arrumos exterior': 'Outside Patio Storage',
}

IGNORE_HEADINGS = {'owner', 'tenant'}

_ws = re.compile(r"\s+")
_dash = re.compile(r"\s*[–—-]\s*")

def norm(s: str) -> str:
    s = s.replace('\u00a0', ' ')
    s = _dash.sub(' - ', s)
    s = _ws.sub(' ', s).strip()
    if s.endswith(':'): s = s[:-1].strip()
    return s

def canonical_room(text: str) -> str | None:
    t = norm(text)
    low = t.lower()
    if low in IGNORE_HEADINGS:
        return None
    # Direct EN match
    for r in ROOMS_CANON:
        if r.lower() == low:
            return r
    # PT mapping
    if low in PT_TO_EN:
        return PT_TO_EN[low]
    return None

rows: list[tuple[str,str,str]] = []
for apt, path in FILES:
    if not path.exists():
        raise FileNotFoundError(path)
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    current_room: str | None = None
    for row in ws.iter_rows(values_only=True):
        # Only look at column A for headers/items, column B and others are owners/notes
        a = row[0]
        if a is None or str(a).strip()=='' :
            continue
        atext = norm(str(a))
        # header if col A matches known heading
        room = canonical_room(atext)
        if room:
            current_room = room
            continue
        # otherwise treat as item under the last seen room
        if current_room is None:
            # default to Hall until first real header appears
            current_room = 'Hall'
        rows.append((apt, current_room, atext))

# Compose migration that rebuilds rows for the four apt types
stamp = time.strftime('%Y%m%d%H%M%S')
out = MIG_DIR / f'{stamp}_checklists_rebuild_v3.sql'
apt_list = sorted({apt for apt,_,_ in rows})
lines = [
    '-- Rebuild checklist_templates from Excel (EN only, PT later)',
    'begin;',
    'set local role postgres;',
    'set statement_timeout = 0;',
    '/* wipe target apt types */',
    'delete from checklist_templates where apartment_type in (' + ','.join(f"'{a}'" for a in apt_list) + ');'
]

order = {a:0 for a in apt_list}
for apt, room, item in rows:
    order[apt]+=1
    i=order[apt]
    lines.append(
        "insert into checklist_templates (apartment_type, room_type, item_description, item_description_pt, order_sequence) "
        f"values ('{apt.replace("'","''")}', '{room.replace("'","''")}', '{item.replace("'","''")}', null, {i});"
    )

lines.append('commit;')
out.write_text('\n'.join(lines)+'\n', encoding='utf-8')
print('Wrote', out)
