from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
import re, time

REPO = Path(__file__).resolve().parents[1].parent
APP = Path(__file__).resolve().parents[1]
MIG_DIR = APP / 'supabase' / 'migrations'

FILES = [
    ('T2', REPO / 'Snag list Inspection - Blank  - T2.xlsx'),
    ('T2+1', REPO / 'Snag list Inspection - Blank  - T2+1.xlsx'),
    ('T3', REPO / 'Snag list Inspection - Blank  - T3.xlsx'),
    ('T3+1', REPO / 'Snag list Inspection - Blank  - T3+1.xlsx'),
]

ROOMS_CANON = [
    'Hall','Kitchen','Dining Room','Living Room','Balcony - Pool side',
    'Suite Bedroom','Suite Bathroom','Bedroom 2','Bathroom 2','Balcony - Back',
    'Outside Patio Storage']
PT_TO_EN = {
    'entrada':'Hall','cozinha':'Kitchen','sala de estar':'Living Room',
    'quarto suite':'Suite Bedroom','casa de banho suite':'Suite Bathroom',
    'quarto 2':'Bedroom 2','casa de banho 2':'Bathroom 2',
    'varanda - pool side':'Balcony - Pool side','varanda - traseira':'Balcony - Back',
    'arrumos exterior':'Outside Patio Storage'
}
IGNORE_HEADINGS = {'owner','tenant'}
_ws = re.compile(r"\s+")
_dash = re.compile(r"\s*[–—-]\s*")

def norm(s: str) -> str:
    s = s.replace('\u00a0',' ')
    s = _dash.sub(' - ', s)
    s = _ws.sub(' ', s).strip()
    if s.endswith(':'): s = s[:-1].strip()
    return s

def canonical_room(txt: str) -> str|None:
    t = norm(txt)
    low = t.lower()
    if low in IGNORE_HEADINGS: return None
    for r in ROOMS_CANON:
        if r.lower()==low: return r
    return PT_TO_EN.get(low)

rows: list[tuple[str,str,str]] = []
for apt, path in FILES:
    if not path.exists():
        raise FileNotFoundError(path)
    ws = load_workbook(path)[0]
    current = None
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if a is None or str(a).strip()=='' :
            continue
        t = norm(str(a))
        room = canonical_room(t)
        if room:
            current = room
            continue
        if current is None:
            current = 'Hall'
        rows.append((apt, current, t))

stamp = time.strftime('%Y%m%d%H%M%S')
out = MIG_DIR / f'{stamp}_checklists_full_reset.sql'
order = {'T2':0,'T2+1':0,'T3':0,'T3+1':0}
lines = [
    '-- FULL RESET of checklist_templates and repopulation from Excel (EN only)',
    'begin;',
    'truncate table checklist_templates restart identity;',
]
for apt, room, item in rows:
    order[apt] = order.get(apt,0) + 1
    i = order[apt]
    lines.append(
        "insert into checklist_templates (apartment_type, room_type, item_description, item_description_pt, order_sequence) "
        f"values ('{apt.replace("'","''")}', '{room.replace("'","''")}', '{item.replace("'","''")}', null, {i});"
    )
lines.append('commit;')
out.write_text('\n'.join(lines)+'\n', encoding='utf-8')
print('Wrote', out)
